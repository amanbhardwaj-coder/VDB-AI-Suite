#!/usr/bin/env python3
"""
MFG → VDB Converter — single-file consolidated app.

Two ways to run:

  Streamlit UI (interactive, with file uploads):
      streamlit run app.py

  Command-line (batch / automation):
      python app.py INPUT.xlsx [INPUT2.xlsx ...] -i IMAGES.csv -o OUTPUT.csv -c Essential

The conversion logic, the Streamlit UI, and the CLI all live here so you
can deploy a single file. The library functions (`convert`, `load_image_map`)
accept either filesystem paths or file-like objects (BytesIO, Streamlit
UploadedFile) — same code path for both contexts.
"""
import argparse
import io
import re
import sys
from decimal import Decimal, ROUND_HALF_UP

import pandas as pd
from openpyxl import load_workbook


# ============================================================================
# CONFIG
# ============================================================================
DEFAULT_COLLECTION = "Essential"             # → Config Field 13 when not overridden
IMG_BASE     = "https://d2dtfeai6qg5ne.cloudfront.net/Smiling_Rocks/"
MAX_IMAGES   = 8                              # Image Url 1..8
SIDE_COLOR, SIDE_CLARITY = "E-F", "SI1"       # 29-May convention
CEN_COLOR,  CEN_CLARITY  = "E-F", "SI1"       # 29-May convention

# Curated remark → (sub-category rule, Config Field 14, Jewelry Style)
# subcat_rule == "jtype" means: use the MFG jewelry type (Ring/Earring/...) as Sub-Category.
# Anything else is used literally for Sub-Category.
REMARK_MAP = {
    "3 STONE":                    ("jtype", "Three Stone",        "Three Stone"),
    "3 STONE BAZEL":              ("jtype", "Three Stone",        "Three Stone"),
    "BAZEL STUD":                 ("jtype", "Solitaire Stud",     "Solitaire Stud"),
    "BEZEL SOLITAIRE":            ("jtype", "Solitaire",          "Solitaire"),
    "CHANNEL FULL ETERNITY BAND": ("jtype", "Full Eternity Band", "Full Eternity Band"),
    "CHANNEL HOOPS":              ("jtype", "In and Out Hoops",   "In and Out Hoops"),
    "CHENNAL SETTING BAND":       ("Band",  "Band",               "Chennal Setting Band"),
    "ENGAGEMENT":                 ("jtype", "Engagement Ring",    "Engagement Ring"),
    "ESSENTIALS":                 ("jtype", "Fashion",            "Fashion"),
    "FLOWER EARRINGS":            ("jtype", "Flower Earring",     "Flower Earring"),
    "HALO RINGS":                 ("jtype", "Halo",               "Halo"),
    "HOOP EARRINGS":              ("jtype", "In and Out Hoops",   "In and Out Hoops"),
    "IN OUT HOOP EARRINGS":       ("jtype", "In and Out Hoops",   "In and Out Hoops"),
    "SOLITAIRE":                  ("jtype", "Solitaire",          "Solitaire"),
    "SOLITAIRE PENDANTS":         ("jtype", "Solitaire Pendant",  "Solitaire Pendant"),
    "SOLITAIRE STUDS":            ("jtype", "Solitaire Stud",     "Solitaire Stud"),
}

SHAPE_MAP = {
    'BAGUETTE':'Baguette','BAGUTTE':'Baguette','BAGUET':'Baguette','BGT':'Baguette','BAG':'Baguette','BUG':'Baguette',
    'RND':'Round','ROUND':'Round','OVL':'Oval','OVAL':'Oval','EMR':'Emerald','EMERALD':'Emerald',
    'PER':'Pear','PEAR':'Pear','CUS':'Cushion','CUSHION':'Cushion','PRN':'Princess','PRINCESS':'Princess',
    'RAD':'Radiant','RADIANT':'Radiant','MAR':'Marquise','MQS':'Marquise','MARQUISE':'Marquise',
    'TRI':'Triangle','ASR':'Asscher','TAP':'Taper','HMN':'Half Moon','HRT':'Heart','HEART':'Heart',
}
COLOR_MAP = {
    'WHSR':'White','WHT':'White','WHITE':'White',
    'BLSR':'Blue','BSSR':'Blue','BLU':'Blue','BLUE':'Blue',
    'PKSR':'Pink','PNK':'Pink','PINK':'Pink',
    'GNSR':'Green','GRSR':'Green','GSR':'Green','TSSR':'Green','GRN':'Green','GREEN':'Green',
    'RDSR':'Red','RDS':'Red','RD':'Red','RED':'Red',
    'YLSR':'Yellow','YEL':'Yellow','YELLOW':'Yellow',
    'RUSR':'Ruby','EMSR':'Emerald','RNBW':'Rainbow',
}
COLOR_NAME = {'W':'White','Y':'Yellow','R':'Rose'}

# Exact 41-column VDB schema (order matters).
COLUMNS = [
    'Master Stock','Stock Number','Jewelry Sub-Category','Jewelry Style','Config Field 13',
    'Config Field 14','Tags','Short Title','Description','Metal',
    'Image Url 1','Image Url 2','Image Url 3','Image Url 4','Image Url 5','Image Url 6','Image Url 7','Image Url 8',
    'Available Metals','Config Field 17','Total Price','Weight','Config field 11','Available Config field 11',
    'Center Carat Weight','Available Center Carat Weight','Center Color','Center Clarity','Shape','Supported Shape Variations',
    'Side Carat Weight','Side Total Stones','Side Color','Side Clarity','Side Stone Shape',
    'Config Field 6','Available Config Field 6','Comments','barcode No','Available Config Field 7','Available Ring Size',
]


def _txt(x):
    if x is None: return ''
    if isinstance(x, float) and x.is_integer(): return str(int(x))
    return str(x).strip()

def _num(x):
    s = _txt(x).replace(',', '')
    try: return float(s)
    except Exception: return 0.0

def _toks(s):
    return re.findall(r'[A-Z0-9]+', _txt(s).upper())

def _r2(v):
    return float(Decimal(str(float(v))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

def _fmt(v):
    v = round(float(v), 4)
    if v == 0: return ''
    if float(v).is_integer(): return str(int(v))
    return f"{v:.4f}".rstrip('0').rstrip('.')

def _fmt2(v):
    v = _r2(v)
    if v == 0: return ''
    if float(v).is_integer(): return str(int(v))
    return f"{v:.2f}".rstrip('0').rstrip('.')

def _fmt2_fixed(v):
    v = _r2(v)
    if v == 0: return ''
    return f"{v:.2f}"

_IMG_CODE_RE = re.compile(r'^([WYR])(\d+)$', re.I)

def load_image_map(path_or_file):
    if not path_or_file:
        return {}, {'accepted': 0, 'skipped': 0, 'skip_samples': {}}
    df = pd.read_csv(path_or_file, dtype=str).fillna('')
    col = df.columns[0]
    tmp = {}
    accepted = 0
    skipped = 0
    skip_samples = {}
    for fn in df[col]:
        fn = fn.strip()
        if not fn or '_' not in fn:
            skipped += 1
            continue
        base, suffix = fn.rsplit('_', 1)
        code = suffix.split('.')[0].strip()
        m = _IMG_CODE_RE.match(code)
        if not m:
            skipped += 1
            skip_samples.setdefault(code, fn)
            continue
        C = m.group(1).upper()
        idx = int(m.group(2))
        tmp.setdefault((base.upper(), C), []).append((idx, fn))
        accepted += 1
    out = {k: [IMG_BASE + fn for _, fn in sorted(v)] for k, v in tmp.items()}
    stats = {'accepted': accepted, 'skipped': skipped, 'skip_samples': skip_samples}
    return out, stats

def _norm_header(x):
    return re.sub(r'[^a-z0-9]+', '', _txt(x).lower())

_HEADER_ALIASES = {
    'no':         ('no',),
    'style':      ('styleno',),
    'desc':       ('description',),
    'pcs':        ('dpcs', 'dpc'),
    'size':       ('dsize',),
    'ct':         ('gmcts', 'gmct'),
    'dc':         ('dc',),
    'srp':        ('srp',),
    'remark':     ('remark',),
    'collection': ('collection',),
}

def _find_header(rows, max_scan=80):
    for r in range(min(max_scan, len(rows))):
        normalized = [_norm_header(c) for c in rows[r]]
        if not all(any(a in normalized for a in _HEADER_ALIASES[f]) for f in ('style', 'desc', 'pcs', 'ct')):
            continue
        idx = {}
        for field, aliases in _HEADER_ALIASES.items():
            for a in aliases:
                if a in normalized:
                    idx[field] = normalized.index(a)
                    break
        return r, idx
    return None, None

def extract(path_or_file):
    wb = load_workbook(path_or_file, read_only=True, data_only=True)
    items = []
    for sheet_name in wb.sheetnames:
        sheet_key = re.sub(r'[^0-9A-Z]+', '', _txt(sheet_name).upper())
        if sheet_key.startswith('18'):
            purity = '18K'
        elif sheet_key.startswith('14'):
            purity = '14K'
        else:
            purity = '18K' if '18' in sheet_key else '14K'
        rows = [list(r) for r in wb[sheet_name].iter_rows(values_only=True)]

        header_row, idx = _find_header(rows)
        if header_row is None:
            continue

        def cell(row, field):
            j = idx.get(field)
            return row[j] if (j is not None and j < len(row)) else None

        cur = None
        for r in range(header_row + 1, len(rows)):
            row = rows[r]
            first = _txt(cell(row, 'no'))
            style = _txt(cell(row, 'style'))
            desc  = _txt(cell(row, 'desc'))

            if first.replace('.0', '').isdigit():
                if cur:
                    if cur.get('base'):
                        items.append(cur)
                m = re.search(r'\bSR[A-Z]+-\d+[A-Z0-9]*\b', style.upper())
                remark_val     = _txt(cell(row, 'remark'))
                collection_val = _txt(cell(row, 'collection'))
                cur = {
                    'sheet': sheet_name, 'purity': purity,
                    'base': (m.group(0) if m else ''),
                    'jtype': '',
                    'remark': remark_val or collection_val,
                    'dc':  _num(cell(row, 'dc')),
                    'srp': _num(cell(row, 'srp')),
                    'total_dia': 0.0,
                    'centers': [], 'sides': [],
                }
            if cur is None:
                continue

            if not cur.get('base'):
                m = re.search(r'\bSR[A-Z]+-\d+[A-Z0-9]*\b', style.upper())
                if m:
                    cur['base'] = m.group(0)

            for cand in (style, desc):
                if cand.upper() in ['RING', 'NECKLACE', 'EARRING', 'BRACELET', 'PENDANT', 'BAND']:
                    cur['jtype'] = cand.title()

            pcs   = _num(cell(row, 'pcs'))
            carat = _num(cell(row, 'ct'))
            tk = _toks(desc)
            if (tk[:1] == ['CS'] or any(c in SHAPE_MAP for c in tk)) and pcs > 0 and carat > 0:
                rec = {
                    'shape': next((SHAPE_MAP[c] for c in tk if c in SHAPE_MAP), ''),
                    'color': next((COLOR_MAP[c] for c in tk if c in COLOR_MAP), ''),
                    'pcs': pcs, 'carat': carat,
                }
                (cur['centers'] if tk[:1] == ['CS'] else cur['sides']).append(rec)

            for j, c in enumerate(row):
                if c and 'Total Dia.' in str(c) and j + 1 < len(row):
                    cur['total_dia'] = _num(row[j + 1])

        if cur:
            if cur.get('base'):
                items.append(cur)
            cur = None
    return items

def _grouped(stones):
    agg, order = {}, []
    for s in stones:
        k = (s['shape'], s['color'])
        if k not in agg:
            agg[k] = {'pcs': 0.0, 'ct': 0.0}
            order.append(k)
        agg[k]['pcs']  += s['pcs']
        agg[k]['ct']   += s['carat']
    lines = []
    for (shape, color) in order:
        v = agg[(shape, color)]
        sc = f"{shape} {color}".strip()
        quality = "(E-F/VS)" if color == "White" else ""
        line = f"{_fmt(v['pcs'])} Pcs | {_fmt2_fixed(v['ct'])} ct | {sc}"
        if quality:
            line += f" {quality}"
        lines.append(re.sub(r'\s+', ' ', line).strip())
    return lines

def _build_comment(centers, sides):
    cl, sl = _grouped(centers), _grouped(sides)
    if centers:
        parts = []
        if cl: parts.append("Center Stone Details\n" + "\n".join(cl))
        if sl: parts.append("Side Stone Details\n" + "\n".join(sl))
        return "\n\n".join(parts)
    return "\n".join(sl)

def convert(inputs, image_map=None, collection=DEFAULT_COLLECTION):
    image_map = image_map or {}
    items = []
    for p in inputs:
        items.extend(extract(p))

    by_base = {}
    for it in items:
        by_base.setdefault(it['base'], {})[it['purity']] = it

    out, unknown = [], set()
    for base in by_base:
        purs = by_base[base]
        _ORD = ('R', 'W', 'Y')
        avail = '#'.join(
            f"{p} {COLOR_NAME[C]} Gold"
            for p in sorted(purs.keys(), key=lambda x: int(x[:2]))
            for C in _ORD
        )

        for purity, it in sorted(purs.items()):
            rk = re.sub(r'\s+', ' ', it['remark'].upper()).strip()
            if rk in REMARK_MAP:
                rule, cf14, style = REMARK_MAP[rk]
                subcat = it['jtype'] if rule == 'jtype' else rule
            else:
                if rk:
                    unknown.add(it['remark'])
                subcat = it['jtype'] or 'Jewelry'
                cf14 = style = it['remark'].title()
            subcat = subcat or 'Jewelry'

            cen_ct   = sum(c['carat'] for c in it['centers'])
            side_ct  = sum(s['carat'] for s in it['sides'])
            side_pcs = sum(s['pcs']   for s in it['sides'])
            weight   = it['total_dia'] if it['total_dia'] > 0 else (cen_ct + side_ct)
            shape = (it['centers'][0]['shape'] if it['centers'] else (it['sides'][0]['shape'] if it['sides'] else 'Round')) or 'Round'
            side_shape = (it['sides'][0]['shape'] if it['sides'] else 'Round') or 'Round'
            comment = _build_comment(it['centers'], it['sides'])
            tags = collection if (cf14 == 'Fashion' or not cf14) else f"{collection},{cf14}"

            for C in ('W', 'Y', 'R'):
                color = COLOR_NAME[C]
                metal = f"{purity} {color} Gold"
                wt = _fmt2_fixed(weight)
                title = re.sub(r'\s+', ' ', f"{collection} {style} {wt}CT {metal} {subcat}").strip()
                urls = image_map.get((base.upper(), C), [])[:MAX_IMAGES]
                imgcols = {f'Image Url {i}': (urls[i-1] if i <= len(urls) else '') for i in range(1, MAX_IMAGES + 1)}

                out.append({
                    'Master Stock':            f"{base}-14KW",
                    'Stock Number':            f"{base}-{purity}{C}",
                    'Jewelry Sub-Category':    subcat,
                    'Jewelry Style':           style,
                    'Config Field 13':         collection,
                    'Config Field 14':         cf14,
                    'Tags':                    tags,
                    'Short Title':             title,
                    'Description':             '',
                    'Metal':                   metal,
                    **imgcols,
                    'Available Metals':        avail,
                    'Total Price':             _fmt(it['dc']),
                    'Config Field 17':         _fmt(it['srp']),
                    'Weight':                  wt,
                    'Center Carat Weight':     _fmt(cen_ct),
                    'Available Center Carat Weight': _fmt2(cen_ct),
                    'Center Color':            CEN_COLOR if cen_ct > 0 else '',
                    'Center Clarity':          CEN_CLARITY if cen_ct > 0 else '',
                    'Shape':                   shape,
                    'Supported Shape Variations': '',
                    'Side Carat Weight':       _fmt(side_ct),
                    'Side Total Stones':       _fmt(side_pcs),
                    'Side Color':              SIDE_COLOR,
                    'Side Clarity':            SIDE_CLARITY,
                    'Side Stone Shape':        side_shape,
                    'Comments':                comment,
                })

    df = pd.DataFrame(out)
    for c in COLUMNS:
        if c not in df.columns:
            df[c] = ''
    df = df[COLUMNS].fillna('')
    return df, unknown

def _run_cli(argv=None):
    ap = argparse.ArgumentParser(description="Convert MFG .xlsx files into VDB-format CSV.")
    ap.add_argument('inputs', nargs='+', help='MFG .xlsx file(s)')
    ap.add_argument('-o', '--output', default='VDB_Output.csv')
    ap.add_argument('-i', '--images', default=None, help='image-filename list CSV (single column)')
    ap.add_argument('-c', '--collection', default=DEFAULT_COLLECTION, help=f"collection name (Config Field 13), default {DEFAULT_COLLECTION!r}")
    a = ap.parse_args(argv)

    imap, stats = load_image_map(a.images)
    if a.images:
        print(f"Image list: accepted {stats['accepted']} W/Y/R+digit files -> {len(imap)} (base,color) groups; skipped {stats['skipped']} non-conforming entries")
        if stats['skip_samples']:
            top = sorted(stats['skip_samples'].items())[:8]
            print("  examples of skipped codes:", ", ".join(f"{c!r} ({fn})" for c, fn in top))

    df, unknown = convert(a.inputs, imap, collection=a.collection)
    df.to_csv(a.output, index=False)
    print(f"Wrote {len(df)} rows -> {a.output}")
    n_styles = df['Stock Number'].str.replace(r'-..K[WYR]$', '', regex=True).nunique()
    print(f"Unique styles: {n_styles}")
    print(f"Rows with at least one image: {(df['Image Url 1'] != '').sum()}/{len(df)}")
    if unknown:
        print("\n[!] Unmapped remarks (defaulted, please add to REMARK_MAP):")
        for u in sorted(unknown):
            print("   ", u)

def _running_under_streamlit():
    import os
    if 'streamlit' not in sys.modules and 'STREAMLIT_SERVER_PORT' not in os.environ:
        return False
    try:
        import logging
        logging.getLogger('streamlit').setLevel(logging.ERROR)
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        return get_script_run_ctx() is not None
    except Exception:
        return False

if _running_under_streamlit():
    import streamlit as st
    from shared.constants import VDB_LOGO_URL

    st.set_page_config(
        page_title="MFG → VDB Converter",
        page_icon=VDB_LOGO_URL,
        layout="wide",
    )

    st.title("MFG → VDB Converter")
    st.caption(
        "Upload a client MFG quotation and the image-filename list, set the "
        "collection name, and download the VDB-format CSV ready to import."
    )

    with st.sidebar:
        st.header("Settings")
        collection = st.text_input(
            "Collection (Config Field 13)",
            value=DEFAULT_COLLECTION,
            help="Shown as Config Field 13 in the output. Change it per MFG file "
                 "(e.g. Essential, Smiling Brides, Fashion).",
        ).strip() or DEFAULT_COLLECTION

        st.markdown("---")
        st.caption(f"**Mapped remarks:** {len(REMARK_MAP)}")
        with st.expander("Show remark → category map"):
            rm_df = pd.DataFrame(
                [(k, v[0], v[1], v[2]) for k, v in REMARK_MAP.items()],
                columns=["MFG Remark", "Sub-Cat rule", "Config Field 14", "Jewelry Style"],
            ).sort_values("MFG Remark").reset_index(drop=True)
            st.dataframe(rm_df, use_container_width=True, hide_index=True)

    col_l, col_r = st.columns(2)
    with col_l:
        st.subheader("1. MFG file(s)")
        mfg_files = st.file_uploader(
            "Client quotation workbook (.xlsx). Multiple files OK.",
            type=["xlsx"],
            accept_multiple_files=True,
            key="mfg",
        )
    with col_r:
        st.subheader("2. Image list (optional)")
        img_file = st.file_uploader(
            "Single-column CSV of image filenames (e.g. SRR-00023WHT_W1.jpg).",
            type=["csv"],
            accept_multiple_files=False,
            key="images",
        )

    disabled = not mfg_files
    run = st.button("Convert →", type="primary", disabled=disabled, use_container_width=True)
    if disabled:
        st.info("Upload at least one MFG .xlsx to enable conversion.")

    if run:
        image_map, img_stats = {}, {'accepted': 0, 'skipped': 0, 'skip_samples': {}}
        if img_file is not None:
            try:
                image_map, img_stats = load_image_map(io.BytesIO(img_file.getvalue()))
            except Exception as e:
                st.error(f"Couldn't read the image list: {e}")
                st.stop()

        try:
            inputs = [io.BytesIO(f.getvalue()) for f in mfg_files]
            with st.spinner("Parsing MFG sheets and building VDB rows…"):
                df, unknown = convert(inputs, image_map, collection=collection)
        except Exception as e:
            st.error(f"Conversion failed: {e}")
            st.stop()

        st.session_state["result_df"]  = df
        st.session_state["unknown"]    = sorted(unknown)
        st.session_state["img_stats"]  = img_stats
        st.session_state["img_used"]   = img_file is not None
        st.session_state["collection"] = collection

    if "result_df" in st.session_state:
        df         = st.session_state["result_df"]
        unknown    = st.session_state["unknown"]
        img_stats  = st.session_state["img_stats"]
        img_used   = st.session_state["img_used"]
        coll_used  = st.session_state["collection"]

        st.markdown("---")
        st.subheader("Results")

        unique_styles = df["Stock Number"].str.replace(r"-..K[WYR]$", "", regex=True).nunique()
        rows_with_img = (df["Image Url 1"] != "").sum()
        cols = st.columns(4)
        cols[0].metric("Rows", f"{len(df):,}")
        cols[1].metric("Unique styles", f"{unique_styles:,}")
        cols[2].metric(
            "Rows with image",
            f"{rows_with_img:,} / {len(df):,}",
            delta=f"{(rows_with_img / len(df) * 100):.1f}%" if len(df) else None,
            delta_color="off",
        )
        cols[3].metric(
            "Image codes accepted",
            f"{img_stats['accepted']:,}" if img_used else "—",
            delta=f"{img_stats['skipped']:,} skipped" if img_used else None,
            delta_color="off",
        )

        if unknown:
            with st.expander(
                f"⚠️ {len(unknown)} MFG remark(s) not in REMARK_MAP — defaulted in output",
                expanded=True,
            ):
                st.write(
                    "These remarks were defaulted (Config Field 14 / Jewelry Style "
                    "set to a title-case copy of the remark). To map them properly, "
                    "edit `REMARK_MAP` near the top of `app.py`."
                )
                st.code("\n".join(unknown))

        if img_used:
            missing_imgs = df[df["Image Url 1"] == ""]
            if not missing_imgs.empty:
                missing_styles = sorted(set(missing_imgs["Stock Number"].str.replace(r"-..K[WYR]$", "", regex=True)))
                with st.expander(f"📷 {len(missing_styles)} style(s) have no image in the list"):
                    st.code(", ".join(missing_styles))

            if img_stats["skip_samples"]:
                with st.expander(
                    f"🪣 {img_stats['skipped']} image-list entries skipped (WG/YG/RG, LS, LF, digits-only, malformed)"
                ):
                    st.caption(
                        "Only W1/Y1/R1-style codes are accepted. Examples of "
                        "skipped codes (one filename per code):"
                    )
                    top = sorted(img_stats["skip_samples"].items())[:25]
                    st.code("\n".join(f"{code!r:14}  {fn}" for code, fn in top))

        st.markdown("**Preview** (first 50 rows)")
        st.dataframe(df.head(50), use_container_width=True, hide_index=True)

        csv_bytes = df.to_csv(index=False).encode("utf-8")
        fname_slug = re.sub(r"[^A-Za-z0-9_-]+", "_", coll_used) or "VDB"
        st.download_button(
            "⬇️ Download VDB CSV",
            data=csv_bytes,
            file_name=f"VDB_{fname_slug}_Output.csv",
            mime="text/csv",
            type="primary",
            use_container_width=True,
        )

elif __name__ == "__main__":
    _run_cli()
